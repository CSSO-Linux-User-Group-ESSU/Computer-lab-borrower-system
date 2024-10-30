from django.shortcuts import render, redirect, get_object_or_404
import django.contrib.messages as messages
from .models import BorrowerInfo
from django.contrib.auth import login, authenticate,logout
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.http import HttpResponse
import subprocess
from .form import UploadFileForm, BorrowerForm
import json
from openpyxl import load_workbook
import cv2


def success_page(request):
    return render(request, "borrower_app/success_page.html")


def upload_and_process_file(request):
    if request.method == 'POST':
        # Handle file upload
        file_form = UploadFileForm(request.POST, request.FILES)

        if file_form.is_valid():
            uploaded_file = request.FILES['files']
            wb = load_workbook(uploaded_file)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                (
                    last_name, first_name, middle_name, projector_qty,
                    led_qty, monitor_qty, keyboard_qty, mouse_qty,
                    cpu_qty, ups_qty, sub_cord_qty, power_cord_qty
                ) = row

                # Skip rows with missing values
                if None in row:
                    continue

                borrower_data = {
                    'last_name': last_name,
                    'first_name': first_name,
                    'middle_name': middle_name,
                    'projector_qty': projector_qty,
                    'led_qty': led_qty,
                    'monitor_qty': monitor_qty,
                    'keyboard_qty': keyboard_qty,
                    'mouse_qty': mouse_qty,
                    'cpu_qty': cpu_qty,
                    'ups_qty': ups_qty,
                    'sub_cord_qty': sub_cord_qty,
                    'power_cord_qty': power_cord_qty,
                }

                borrower_form = BorrowerForm(borrower_data)

                if borrower_form.is_valid():
                    # Save the valid borrower data to the database
                    borrower_form.save()
                else:
                    # Print or log form errors for debugging purposes
                    print(f"Form errors for row: {borrower_form.errors}")

            messages.success(request, "File processed successfully!")
            return redirect('success_page')
        else:
            messages.error(request, "Invalid file upload. Please try again.")

    else:
        file_form = UploadFileForm()

    # Render the form for file upload
    return render(request, 'borrower_app/upload_and_process_file.html', {'file_form': file_form})


def manual_input(request):
    return render(request, 'borrower_app/form.html')


# Create your views here.
def logins(request):
    return render(request, 'borrower_app/login.html')


def sign_in(request):
    return render(request, 'borrower_app/sign_in.html')

def log_out(request):
    logout(request)
    return redirect('/')


def control_panel(request):
    if request.method == "POST":
        name = request.POST.get("username")  # Safely retrieve the username
        password = request.POST.get("password")  # Safely retrieve the password

        # Authenticate the user
        user = authenticate(request, username=name, password=password)

        if user is not None:
            login(request, user)  # Log the user in
            return redirect('home')  # Redirect to the home page after successful login
        else:
            messages.error(request, "Wrong username or password! Please try again.")
            return redirect('login')  # Redirect to the login page on error

    return render(request, 'borrower_app/login.html')  # Render the login page for GET requests

@login_required
def home(request):
    return render(request, 'borrower_app/home.html')


def return_items(request):
    return render(request, 'borrower_app/return_items.html')


def pending_items(request):
    borrowers = BorrowerInfo.objects.all()
    borrower_data = []
    for borrower in borrowers:
        total = borrower.total_quantity()
        borrower_data.append({
            'id': borrower.id,  # Ensure borrower.id is passed here
            'last_name': borrower.last_name,
            'first_name': borrower.first_name,
            'middle_name': borrower.middle_name,
            'projector_qty': borrower.projector_qty,
            'led_qty': borrower.led_qty,
            'monitor_qty': borrower.monitor_qty,
            'keyboard_qty': borrower.keyboard_qty,
            'mouse_qty': borrower.mouse_qty,
            'cpu_qty': borrower.cpu_qty,
            'ups_qty': borrower.ups_qty,
            'sub_cord_qty': borrower.sub_cord_qty,
            'power_cord_qty': borrower.power_cord_qty,
            'total': total,
            # add date
        })
    return render(request, 'borrower_app/pending_items.html', {'borrowers': borrower_data})



def delete_borrower(request, borrower_id):
    borrower = get_object_or_404(BorrowerInfo, id=borrower_id)

    if request.method == "POST":
        borrower.delete()
        messages.success(request, "Borrower deleted successfully!")
        return redirect('pending_items')

    return render(request, 'borrower_app/confirm_delete.html', {'borrower': borrower})



def delete_borrowers(request):
    if request.method == "POST":

        data = json.loads(request.body)
        borrower_ids = data.get('borrower_ids', [])

        # Delete the borrowers with the provided IDs
        BorrowerInfo.objects.filter(id__in=borrower_ids).delete()

        # Return a success response
        return JsonResponse({'success': True})
    return JsonResponse({'success': False})


def edit_borrower(request, borrower_id):
    borrower = get_object_or_404(BorrowerInfo, id=borrower_id)

    if request.method == "POST":
        # Update borrower fields with data from the form
        borrower.last_name = request.POST.get('last_name')
        borrower.first_name = request.POST.get('first_name')
        borrower.middle_name = request.POST.get('middle_name')
        borrower.projector_qty = request.POST.get('projector_qty')
        borrower.led_qty = request.POST.get('led_qty')
        borrower.monitor_qty = request.POST.get('monitor_qty')
        borrower.keyboard_qty = request.POST.get('keyboard_qty')
        borrower.mouse_qty = request.POST.get('mouse_qty')
        borrower.cpu_qty = request.POST.get('cpu_qty')
        borrower.ups_qty = request.POST.get('ups_qty')
        borrower.sub_cord_qty = request.POST.get('sub_cord_qty')
        borrower.power_cord_qty = request.POST.get('power_cord_qty')
        borrower.save()
        return redirect('pending_items')

    return render(request, 'borrower_app/edit_borrower.html', {'borrower': borrower})


def scan_printer(request):
    has_printer = subprocess.check_output('lpstat -p', shell=True)
    print(has_printer,'has_printer')
    if b'disabled' in has_printer:
        return HttpResponse('No Printer')
    else:
        return HttpResponse('Has Printer')


def scan_paper(request):
    image_path = 'HTR/scannedImages/scanned_form.png'
    with open(image_path, 'wb') as f:
        is_scan = subprocess.run('scanimage --format=png --mode Color --resolution 600 --brightness 50 --contrast 50', stdout=f, shell=True)
        if is_scan.returncode == 0:

            # Cropping the scanned image
            image = cv2.imread(image_path)
            h, w, _ = image.shape
            image = image[:,:w//2]
            cv2.imwrite(image_path, image)
            
            # Doing now the actual scanning
            scan = subprocess.run(f'python3 HTR/ocr.py {image_path}', shell=True)
            if scan.returncode == 0:
                return HttpResponse('Scanned')
            else:
                return HttpResponse('Not Scanned')
        else:
            return HttpResponse('Not Scanned')